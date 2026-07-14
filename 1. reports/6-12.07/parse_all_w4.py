#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse all W4 data. Output: parsed_w4.json"""
import json, re
from collections import defaultdict
import openpyxl

BASE = '1. reports/6-12.07'
OUT = {}

# ── DIRECT MSK ──
# col0=campaign, col44=spend, col45=clicks (standard report, rows 1-5 = metadata)
wb = openpyxl.load_workbook(f'{BASE}/2026-07-14_16-19-28_sk-tu-msk-516295-pguz.xlsx', data_only=True)
ws = wb.active
SKIP_MSK = ['Итого', 'Клиент:', 'Отчет:', 'Период:', 'Название кампании']
msk = defaultdict(lambda: {'spend': 0.0, 'clicks': 0, 'impressions': 0})
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[0]).strip() if row[0] else ''
    if not name or any(k in name for k in SKIP_MSK):
        continue
    try:
        s = float(row[44]) if row[44] is not None and str(row[44]) != '-' else 0.0
        c = int(float(row[45])) if row[45] is not None and str(row[45]) != '-' else 0
    except (ValueError, TypeError):
        continue
    msk[name]['spend'] += s
    msk[name]['clicks'] += c

t_msk_s = sum(d['spend'] for d in msk.values())
t_msk_c = sum(d['clicks'] for d in msk.values())
t_msk_i = sum(d['impressions'] for d in msk.values())
OUT['direct_msk'] = {'campaigns': {k: v for k, v in sorted(msk.items()) if v['spend'] > 0}, 'total_spend': round(t_msk_s,2), 'total_clicks': t_msk_c, 'total_impressions': t_msk_i, 'campaign_count': len([k for k,v in msk.items() if v['spend'] > 0])}

# ── DIRECT SPB ──
# Headers row 5: col0=campaign, col33=spend, col34=clicks
wb2 = openpyxl.load_workbook(f'{BASE}/2026-07-14_16-21-11_sk-tu-ru-424188-lwwn.xlsx', data_only=True)
ws2 = wb2.active
spb = defaultdict(lambda: {'spend': 0.0, 'clicks': 0, 'impressions': 0})
for row in ws2.iter_rows(min_row=6, max_row=ws2.max_row, values_only=True):
    name = str(row[0]).strip() if row[0] else ''
    if not name or name == 'Итого':
        continue
    try:
        s = float(row[33]) if row[33] is not None and str(row[33]) != '-' else 0.0
        c = int(float(row[34])) if row[34] is not None and str(row[34]) != '-' else 0
    except (ValueError, TypeError):
        continue
    spb[name]['spend'] += s
    spb[name]['clicks'] += c

t_spb_s = sum(d['spend'] for d in spb.values())
t_spb_c = sum(d['clicks'] for d in spb.values())
OUT['direct_spb'] = {'campaigns': {k: v for k, v in sorted(spb.items())}, 'total_spend': round(t_spb_s,2), 'total_clicks': t_spb_c, 'campaign_count': len(spb)}

# ── CRM ──
with open(f'{BASE}/CRM 6-12.txt', 'r', encoding='utf-8') as f:
    text = f.read()
# Parse ALL JSON blocks (CRM export has multiple arrays)
crm_all = []
for m in re.finditer(r'```(?:json)?\s*(\[.*?\])\s*```', text, re.DOTALL):
    try:
        block = json.loads(m.group(1))
        crm_all.extend(block)
    except:
        pass
# Also try raw arrays outside code blocks
for m in re.finditer(r'(?<![`\w])\[\s*\{[^]]+\}\s*\]', text, re.DOTALL):
    try:
        block = json.loads(m.group(0))
        crm_all.extend(block)
    except:
        pass
# Dedupe by title+phone+status (some deals appear in multiple exports)
seen = set()
crm_data = []
for r in crm_all:
    key = (r.get('title',''), r.get('phone',''), r.get('status',''))
    if key not in seen:
        seen.add(key)
        crm_data.append(r)

TARGET = ['Квалифицирован','КП направлено','Встреча проведена','Дожим','Договор направлен','Передан на расчет','Встреча назначена','ОС по КП получено','Встреча подтверждена','Смета рассчитана','ДОГОВОР ПОДПИСАН']

statuses = defaultdict(int); target_ct = 0; target_list = []
for r in crm_data:
    st = r.get('status','').strip()
    statuses[st] += 1
    if any(t.lower() in st.lower() for t in TARGET):
        target_ct += 1
        target_list.append({'source': r.get('source',''), 'title': r.get('title',''), 'status': st, 'date': r.get('date',''), 'tag': r.get('tag','')})

OUT['crm'] = {'total': len(crm_data), 'target_leads': target_ct, 'status_counts': dict(statuses), 'target_details': target_list}

# ── WEBMASTER ──
WM = {
    'msk_pages': f'{BASE}/msk.sk-tu.ru_b339261c52cabee5e1510dec.xlsx',
    'msk_queries': f'{BASE}/msk.sk-tu.ru_c9464c2f84604c94a39a9888.xlsx',
    'spb_queries': f'{BASE}/sk-tu.ru_1894984520dd5f041f52dc58.xlsx',
    'spb_pages': f'{BASE}/sk-tu.ru_f55840f0a957a92fb6569b6e.xlsx',
}

wm_out = {}
for key, fp in WM.items():
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value else '' for c in ws[1]]
    # Find column indices
    def ci(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n.lower() in h.lower():
                    return i
        return None
    imp_i = ci('impressions') or 2
    clk_i = ci('clicks') or 3
    ctr_i = ci('ctr') or 4
    pos_i = ci('position') or 5
    
    total_imp = 0; total_clk = 0
    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        nm = str(row[0]).strip() if row[0] else ''
        if not nm or 'ALL_QUERIES' in nm.upper():
            continue
        try:
            imp = int(float(row[imp_i])) if row[imp_i] is not None else 0
            clk = int(float(row[clk_i])) if row[clk_i] is not None else 0
            ctr = float(row[ctr_i]) if row[ctr_i] is not None else 0
            pos = float(row[pos_i]) if row[pos_i] is not None else 0
        except:
            continue
        total_imp += imp; total_clk += clk
        if clk > 0:
            entries.append({'name': nm, 'impressions': imp, 'clicks': clk, 'ctr': ctr, 'position': pos})
    entries.sort(key=lambda x: -x['clicks'])
    wm_out[key] = {'total_impressions': total_imp, 'total_clicks': total_clk, 'ctr': round(total_clk/total_imp*100,2) if total_imp>0 else 0, 'top': entries[:20]}

OUT['webmaster'] = wm_out

# ── SUMMARY ──
OUT['summary'] = {
    'period': '06.07-12.07.2026',
    'week': 4,
    'direct_total_spend': round(t_msk_s + t_spb_s, 2),
    'direct_total_clicks': t_msk_c + t_spb_c,
    'direct_msk_spend': round(t_msk_s,2),
    'direct_msk_clicks': t_msk_c,
    'direct_spb_spend': round(t_spb_s,2),
    'direct_spb_clicks': t_spb_c,
    'webmaster_total_impressions': sum(v['total_impressions'] for k,v in wm_out.items() if 'pages' in k),
    'webmaster_total_clicks': sum(v['total_clicks'] for k,v in wm_out.items() if 'pages' in k),
    'crm_total': len(crm_data),
    'crm_target_leads': target_ct
}

with open(f'{BASE}/parsed_w4.json', 'w', encoding='utf-8') as f:
    json.dump(OUT, f, ensure_ascii=False, indent=2)

# Print summary (safe chars only)
print("=== WEEK 4 PARSING COMPLETE ===")
print(f"Direct MSK: {t_msk_s:,.2f} rub | {t_msk_c:,} clicks | {t_msk_i:,} imp | {len(msk)} campaigns")
print(f"Direct SPB: {t_spb_s:,.2f} rub | {t_spb_c:,} clicks | {len(spb)} campaigns")
print(f"Direct TOTAL: {t_msk_s+t_spb_s:,.2f} rub | {t_msk_c+t_spb_c:,} clicks")
print(f"CRM: {len(crm_data)} records | {target_ct} target leads")
for k, v in wm_out.items():
    print(f"Webmaster {k}: {v['total_impressions']:,} imp | {v['total_clicks']:,} clicks | CTR {v['ctr']}%")
print(f"\nSaved to {BASE}/parsed_w4.json")
